"""
Excel比较工具 - 后端API
核心算法在服务器端执行，保护代码安全
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import xlsxwriter
import io
import base64
import tempfile
import os

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

app = Flask(__name__)
CORS(app)

API_KEY = "your-secret-api-key-here"

def require_api_key(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        auth_header = request.headers.get('Authorization', '')
        if auth_header != f'Bearer {API_KEY}':
            return jsonify({'error': 'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated

def parse_excel(file_storage):
    """解析Excel文件，返回数据结构"""
    if HAS_OPENPYXL:
        wb = openpyxl.load_workbook(file_storage, data_only=True)
        result = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append([str(cell) if cell is not None else '' for cell in row])
            result[sheet_name] = data
        return result
    else:
        return {'error': 'openpyxl not installed'}

def get_chapter_from_row(row, col_index):
    """从行中提取章节号"""
    if not isinstance(row, list) or len(row) <= col_index:
        return None
    import re
    text = str(row[col_index]).strip()
    match = re.match(r'^(\d+(?:\.\d+)*)', text)
    return match.group(1) if match else None

def is_level1_chapter(chapter_str):
    """判断是否为一级章节"""
    if not chapter_str:
        return False
    import re
    return bool(re.match(r'^\d+$', chapter_str.strip()))

def group_by_chapters(data, chapter_col):
    """按章节分组数据"""
    groups = []
    current_group = None
    
    for i, row in enumerate(data):
        chapter = get_chapter_from_row(row, chapter_col)
        
        if chapter and is_level1_chapter(chapter):
            if current_group and current_group['rows']:
                groups.append(current_group)
            current_group = {'start_idx': i, 'rows': [{'row': row, 'idx': i}]}
        else:
            if not current_group:
                current_group = {'start_idx': i, 'rows': []}
            current_group['rows'].append({'row': row, 'idx': i})
    
    if current_group and current_group['rows']:
        groups.append(current_group)
    
    return groups

def rows_equal(row1, row2):
    """比较两行是否相等"""
    if not row1 and not row2:
        return True
    if not row1 or not row2:
        return False
    
    max_len = max(len(row1), len(row2))
    for i in range(max_len):
        v1 = str(row1[i] if i < len(row1) else '').strip()
        v2 = str(row2[i] if i < len(row2) else '').strip()
        if v1 != v2:
            return False
    return True

def compare_sheets_core(data1, data2, chapter_col=0, start_row1=0, start_row2=0):
    """
    核心比较算法 - 这是你要保护的代码
    
    参数:
        data1: 旧文件数据 {sheet_name: [[row1], [row2], ...]}
        data2: 新文件数据
        chapter_col: 章节列索引
        start_row1: 文件1起始行
        start_row2: 文件2起始行
    
    返回:
        {
            'differences': [...],
            'stats': {'added': N, 'deleted': N, 'modified': N, 'identical': N}
        }
    """
    all_sheets = set(data1.keys()) | set(data2.keys())
    differences = []
    stats = {'added': 0, 'deleted': 0, 'modified': 0, 'identical': 0}
    
    for sheet_name in all_sheets:
        raw1 = data1.get(sheet_name, [])
        raw2 = data2.get(sheet_name, [])
        
        sheet_data1 = raw1[start_row1:]
        sheet_data2 = raw2[start_row2:]
        
        groups1 = group_by_chapters(sheet_data1, chapter_col)
        groups2 = group_by_chapters(sheet_data2, chapter_col)
        
        for g in range(max(len(groups1), len(groups2))):
            group1 = groups1[g] if g < len(groups1) else {'start_idx': float('inf'), 'rows': []}
            group2 = groups2[g] if g < len(groups2) else {'start_idx': float('inf'), 'rows': []}
            
            chapter_map1 = {}
            for item in group1['rows']:
                ch = get_chapter_from_row(item['row'], chapter_col)
                if ch:
                    if ch not in chapter_map1:
                        chapter_map1[ch] = []
                    chapter_map1[ch].append(item)
            
            chapter_map2 = {}
            for item in group2['rows']:
                ch = get_chapter_from_row(item['row'], chapter_col)
                if ch:
                    if ch not in chapter_map2:
                        chapter_map2[ch] = []
                    chapter_map2[ch].append(item)
            
            no_chapter1 = [item for item in group1['rows'] if not get_chapter_from_row(item['row'], chapter_col)]
            no_chapter2 = [item for item in group2['rows'] if not get_chapter_from_row(item['row'], chapter_col)]
            
            for item2 in group2['rows']:
                row2 = item2['row']
                idx2 = item2['idx']
                chapter2 = get_chapter_from_row(row2, chapter_col)
                
                if chapter2:
                    old_rows = chapter_map1.get(chapter2)
                    new_rows = chapter_map2.get(chapter2, [])
                    
                    if not old_rows:
                        for col, val in enumerate(row2):
                            if str(val).strip():
                                differences.append({
                                    'type': 'added',
                                    'sheet': sheet_name,
                                    'row': idx2 + start_row2,
                                    'col': col,
                                    'old': '',
                                    'new': str(val)
                                })
                                stats['added'] += 1
                    else:
                        pos = -1
                        for j, nr in enumerate(new_rows):
                            if nr['idx'] == idx2:
                                pos = j
                                break
                        
                        if pos >= 0 and pos < len(old_rows):
                            match = old_rows[pos]
                            if not rows_equal(match['row'], row2):
                                for col in range(max(len(match['row']), len(row2))):
                                    v1 = str(match['row'][col] if col < len(match['row']) else '').strip()
                                    v2 = str(row2[col] if col < len(row2) else '').strip()
                                    
                                    if v1 != v2:
                                        diff_type = 'modified' if v1 and v2 else ('added' if v2 else 'deleted')
                                        differences.append({
                                            'type': diff_type,
                                            'sheet': sheet_name,
                                            'row': idx2 + start_row2,
                                            'col': col,
                                            'old': v1,
                                            'new': v2
                                        })
                                        stats[diff_type] += 1
                                    else:
                                        stats['identical'] += 1
                else:
                    nc_idx = next((i for i, item in enumerate(no_chapter2) if item['idx'] == idx2), -1)
                    match = no_chapter1[nc_idx] if nc_idx >= 0 and nc_idx < len(no_chapter1) else None
                    
                    if not match:
                        for col, val in enumerate(row2):
                            if str(val).strip():
                                differences.append({
                                    'type': 'added',
                                    'sheet': sheet_name,
                                    'row': idx2 + start_row2,
                                    'col': col,
                                    'old': '',
                                    'new': str(val)
                                })
                                stats['added'] += 1
            
            for chapter, old_rows in chapter_map1.items():
                new_rows = chapter_map2.get(chapter, [])
                if len(old_rows) > len(new_rows):
                    deleted_count = len(old_rows) - len(new_rows)
                    for k in range(deleted_count):
                        del_item = old_rows[len(old_rows) - 1 - k]
                        for col, val in enumerate(del_item['row']):
                            if str(val).strip():
                                differences.append({
                                    'type': 'deleted',
                                    'sheet': sheet_name,
                                    'row': del_item['idx'] + start_row1,
                                    'col': col,
                                    'old': str(val),
                                    'new': ''
                                })
                                stats['deleted'] += 1
    
    return {'differences': differences, 'stats': stats}

@app.route('/api/compare', methods=['POST'])
@require_api_key
def compare_files():
    """
    比较两个Excel文件
    
    请求:
        - file1: 旧版本文件
        - file2: 新版本文件
        - chapter_col: 章节列索引（可选，默认0）
        - start_row1: 文件1起始行（可选，默认0）
        - start_row2: 文件2起始行（可选，默认0）
    
    响应:
        {
            'success': true,
            'differences': [...],
            'stats': {'added': N, 'deleted': N, 'modified': N, 'identical': N}
        }
    """
    try:
        if 'file1' not in request.files or 'file2' not in request.files:
            return jsonify({'error': 'Missing files'}), 400
        
        file1 = request.files['file1']
        file2 = request.files['file2']
        
        chapter_col = int(request.form.get('chapter_col', 0))
        start_row1 = int(request.form.get('start_row1', 0))
        start_row2 = int(request.form.get('start_row2', 0))
        
        data1 = parse_excel(file1)
        data2 = parse_excel(file2)
        
        if 'error' in data1 or 'error' in data2:
            return jsonify({'error': 'Failed to parse Excel files'}), 400
        
        result = compare_sheets_core(data1, data2, chapter_col, start_row1, start_row2)
        
        return jsonify({
            'success': True,
            'differences': result['differences'],
            'stats': result['stats']
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export', methods=['POST'])
@require_api_key
def export_result():
    """
    导出比较结果为Excel文件
    
    返回生成的Excel文件（带差异标记）
    """
    try:
        data = request.json
        differences = data.get('differences', [])
        stats = data.get('stats', {})
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        
        worksheet = workbook.add_worksheet('比较结果')
        
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#6366F1',
            'font_color': 'white'
        })
        added_format = workbook.add_format({'bg_color': '#90EE90'})
        deleted_format = workbook.add_format({'bg_color': '#FF0000', 'font_color': 'white'})
        modified_format = workbook.add_format({'bg_color': '#FFFF00'})
        
        worksheet.write(0, 0, '类型', header_format)
        worksheet.write(0, 1, '工作表', header_format)
        worksheet.write(0, 2, '位置', header_format)
        worksheet.write(0, 3, '旧值', header_format)
        worksheet.write(0, 4, '新值', header_format)
        
        for i, diff in enumerate(differences, 1):
            fmt = {'added': added_format, 'deleted': deleted_format, 'modified': modified_format}.get(diff['type'])
            worksheet.write(i, 0, diff['type'], fmt)
            worksheet.write(i, 1, diff['sheet'])
            worksheet.write(i, 2, f"{diff['row']},{diff['col']}")
            worksheet.write(i, 3, diff['old'])
            worksheet.write(i, 4, diff['new'])
        
        stats_sheet = workbook.add_worksheet('统计')
        stats_sheet.write(0, 0, '统计项', header_format)
        stats_sheet.write(0, 1, '数量', header_format)
        stats_sheet.write(1, 0, '新增')
        stats_sheet.write(1, 1, stats.get('added', 0))
        stats_sheet.write(2, 0, '删除')
        stats_sheet.write(2, 1, stats.get('deleted', 0))
        stats_sheet.write(3, 0, '修改')
        stats_sheet.write(3, 1, stats.get('modified', 0))
        
        workbook.close()
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='comparison_result.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
